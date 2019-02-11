import wx
import wx.grid as gridlib
from openpyxl import load_workbook
import xlrd
import csv
import os
import datetime

class MyWin(wx.Frame):

    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title = title, \
                                style = wx.DEFAULT_FRAME_STYLE ^ wx.RESIZE_BORDER ^ wx.MAXIMIZE_BOX)

        self.fileCnt    = 0
        self.files      = []
        self.initAppAndUi()
        self.rows       =[]
        self.sortIndex  = 0


    def initAppAndUi(self):
        # self.frame   = wx.Frame(self, -1, 'win.py', \
        #                         style = wx.DEFAULT_FRAME_STYLE ^ wx.RESIZE_BORDER ^ wx.MAXIMIZE_BOX)
        self.SetSize(0, 0, 850, 600)
        self.panel   = wx.Panel(self, wx.ID_ANY)

        self.list = wx.ListCtrl(self.panel, wx.ID_ANY, (15, 15), (550, 150), style=wx.LC_REPORT)
        self.list.InsertColumn(0, 'File Name', width = 370)
        self.list.InsertColumn(1, 'Organization', width = 170)
        # self.list.Bind(wx.EVT_LIST_ITEM_SELECTED, self.on_list_clicked)

        self.btnAdd  = wx.Button(self.panel, wx.ID_ANY, 'Add', (730, 15))
        self.btnAdd.Bind(wx.EVT_BUTTON, self.on_btnAdd_clicked)
        self.btnDel  = wx.Button(self.panel, wx.ID_ANY, 'Remove', (730, 50))
        self.btnDel.Bind(wx.EVT_BUTTON, self.on_btnRemove_clicked)

        # self.choice = wx.Choice(self.panel, wx.ID_ANY, (590, 100), (227, 100))
        # self.choice.Bind(wx.EVT_CHOICE, self.on_choice_item_selected)
        # self.choice.Append('Organization-1 "Received" Money CSV')
        # self.choice.Append('Organization-1 "Sent" Money CSV')
        # self.choice.Append('Organization-2 "Received" Money CSV')
        # self.choice.Append('Organization-2 "Sent" Money CSV')
        # self.choice.Append('Organization-3 "Received & Sent" Money CSV')
        # self.choice.Append('Organization-4 "Received & Sent" Money CSV')

        self.btnGen  = wx.Button(self.panel, wx.ID_ANY, 'Merge', (730, 140))
        self.btnGen.Bind(wx.EVT_BUTTON, self.on_btnMerge_clicked)

        self.grid    = gridlib.Grid(self.panel, wx.ID_ANY, (15, 180), (800, 330))
        self.grid.CreateGrid(10, 7)
        # grid.UseNativeColHeader(True)
        # grid.SetUseNativeColLabels(True)
        self.grid.SetColMinimalWidth(0, 100)
        self.grid.SetColSize(0, 100)
        self.grid.SetColMinimalWidth(1, 130)
        self.grid.SetColSize(1, 130)
        self.grid.SetColMinimalWidth(2, 60)
        self.grid.SetColSize(2, 60)
        self.grid.SetColMinimalWidth(3, 80)
        self.grid.SetColSize(3, 80)
        self.grid.SetColMinimalWidth(4, 80)
        self.grid.SetColSize(4, 80)
        self.grid.SetColMinimalWidth(5, 100)
        self.grid.SetColSize(5, 100)
        self.grid.SetColMinimalWidth(6, 150)
        self.grid.SetColSize(6, 150)

        self.grid.SetColLabelValue(0, "Organization")
        self.grid.SetColLabelValue(1, "Date")
        self.grid.SetColLabelValue(2, "Currency")
        self.grid.SetColLabelValue(3, "Received")
        self.grid.SetColLabelValue(4, "Sent")
        self.grid.SetColLabelValue(5, "Type")
        self.grid.SetColLabelValue(6, "Description")
        self.grid.SetDefaultCellAlignment(wx.ALIGN_CENTRE,wx.ALIGN_CENTRE)

        # self.btnSort    = wx.Button(self.panel, wx.ID_ANY, 'Sort', (300, 520))
        # self.btnSort.Bind(wx.EVT_BUTTON, self.on_btnSort_clicked)

        self.cmbSort = wx.Choice(self.panel, wx.ID_ANY, (300, 520))
        self.cmbSort.Bind(wx.EVT_CHOICE, self.on_cmbSort_item_selected)
        self.cmbSort.Append('Organization Asc')
        self.cmbSort.Append('Organization Desc')
        self.cmbSort.Append('Date Asc')
        self.cmbSort.Append('Date Desc')
        self.cmbSort.Append('Currency Asc')
        self.cmbSort.Append('Currency Desc')
        self.cmbSort.Append('Received Asc')
        self.cmbSort.Append('Received Desc')
        self.cmbSort.Append('Sent Asc')
        self.cmbSort.Append('Sent Desc')
        self.cmbSort.Append('Type Asc')
        self.cmbSort.Append('Type Desc')
        self.cmbSort.Append('Description Asc')
        self.cmbSort.Append('Description Desc')

        self.btnExport   = wx.Button(self.panel, wx.ID_ANY, 'Export', (450, 520))
        self.btnExport.Bind(wx.EVT_BUTTON, self.on_btnExport_clicked)

        self.Show()
        self.Centre()

        self.Bind(wx.EVT_CLOSE, self.on_close)

    def on_close(self, event):
        self.Destroy()

    # def on_list_clicked(self, event):
        # index   = event.GetIndex()
        # org     = self.list.GetItem(index, 1).GetText()
        # self.choice.SetLabelText(org)

    def xldate_to_datetime(self, xldate):
        temp = datetime.datetime(1900, 1, 1)
        delta = datetime.timedelta(days=xldate)
        return temp + delta

    # def sort_data(self, data):
    #     cnt     = len(data)
    #     for i in range(cnt - 1):
    #         for

    def sort_field(self, val):
        if (self.sortIndex == 3 or self.sortIndex == 4):
            val = str(val[self.sortIndex])
            if (val.strip() == ''):
                return 0

            return float(str(val))

        return val[self.sortIndex]

    def on_btnAdd_clicked(self, event):
        with wx.FileDialog(self, "Open XYZ file", wildcard="Supported Files (*.csv;*.xls;*.xlsx)|*.csv;*.xls;*.xlsx|" \
                                                        "Excel Files (*.xls;*.xlsx)|*.xls;*.xlsx|"\
                                                        "Comma-Separated Values Files (*.csv)|*.csv",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return ''

            fileName    = fileDialog.GetPath()
            fileCnt     = self.list.GetItemCount()
            ifAdded     = False
            for i in range(0, fileCnt):
                fileName1   = self.list.GetItem(i, 0).GetText()
                if (fileName == fileName1):
                    wx.MessageBox("This file is already appended", "Ops", wx.OK | wx.ICON_WARNING)
                    ifAdded = True

            if (ifAdded == False):
                self.list.InsertItem(self.fileCnt, fileName)
                self.fileCnt    += 1


    def on_btnRemove_clicked(self, event):
        if (self.list.GetSelectedItemCount() == 0):
            wx.MessageBox("Select file to remove", "Ops", wx.OK | wx.ICON_WARNING)
            return

        button  = wx.MessageBox("This file is already appended", "Ops", wx.YES | wx.NO | wx.ICON_QUESTION)
        if (button == wx.YES):
            self.list.DeleteItem(self.list.GetFirstSelected())


    def on_choice_item_selected(self, event):
        if (self.list.GetSelectedItemCount() == 0):
            wx.MessageBox("Select file in list", "Ops", wx.OK | wx.ICON_WARNING)
            return
        self.list.SetItem(self.list.GetFirstSelected(), 1, self.choice.GetString(self.choice.GetSelection()))


    def on_btnMerge_clicked(self, event):
        fileCnt     = self.list.GetItemCount()
        for i in range(0, fileCnt):
            fileName1    = self.list.GetItem(i, 0).GetText()
            fileName    = os.path.basename(fileName1);
            fileName, extension = os.path.splitext(fileName)
            ids         = fileName.split('-')
            if (len(ids) < 2):
                wx.MessageBox("Filename is invalid: \"" + fileName1 + "\"", "Ops", wx.OK | wx.ICON_WARNING)

            ids[1]      = ids[1].lower()
            if (not(ids[1] == 'received' or ids[1] == 'sent' or ids[1] == 'all')):
                wx.MessageBox("Filename is invalid: \"" + fileName1 + "\"", "Ops", wx.OK | wx.ICON_WARNING)
                return

        numRows = self.grid.GetNumberRows()
        if (numRows > 0):
            self.grid.DeleteRows(numRows = numRows)
            # self.grid.ClearGrid()

        rIdx = 0

        try:
            for i in range(0, fileCnt):
                fileName    = self.list.GetItem(i, 0).GetText()
                fileName1   = os.path.basename(fileName);
                fileName1, extension = os.path.splitext(fileName1)
                ids         = fileName1.split('-')
                org         = ids[0]
                trans       = ids[1]

                # self.choice.Append('Organization-1 "Received" Money CSV')
                # self.choice.Append('Organization-1 "Sent" Money CSV')
                # self.choice.Append('Organization-2 "Received" Money CSV')
                # self.choice.Append('Organization-2 "Sent" Money CSV')
                # self.choice.Append('Organization-3 "Received & Sent" Money CSV')
                # self.choice.Append('Organization-4 "Received & Sent" Money CSV')
                if (trans == 'received'):
                    dateCol = 0
                    curCol  = 1
                    receCol = 2
                    sentCol = -1
                    typeCol = -1
                    descCol = 3
                elif (trans == 'sent'):
                    dateCol = 0
                    curCol  = 1
                    receCol = -1
                    sentCol = 2
                    typeCol = -1
                    descCol = 3
                elif (trans == 'all'):
                    dateCol = 0
                    curCol  = 1
                    receCol = -2
                    sentCol = -2
                    typeCol = 3
                    descCol = 4

                f, extension = os.path.splitext(fileName)
                print(fileName1,trans)

                if (extension == '.xlsx'):
                    wb      = load_workbook(fileName)
                    sheet   = wb.active
                    rCnt    = sheet.max_row
                    for i in range(2, rCnt + 1):
                        self.grid.InsertRows(rIdx, 1)
                        self.grid.SetCellValue(rIdx, 0, org)
                        val = str(sheet.cell(row=i, column=dateCol + 1).value)
                        self.grid.SetCellValue(rIdx, 1, val)
                        # ///////////////
                        if (curCol != -1):
                            val = str(sheet.cell(row=i, column=curCol + 1).value)
                        else:
                            val = ''
                        self.grid.SetCellValue(rIdx, 2, val)

                        # ///////////////
                        if (typeCol != -1):
                            type = val = str(sheet.cell(row=i, column=typeCol + 1).value)
                        else:
                            type = val = ''
                        self.grid.SetCellValue(rIdx, 5, val)

                        # ///////////////
                        print(receCol, sentCol)
                        if (receCol == -2):
                            if (type == 'Received'):
                                val = str(sheet.cell(row=i, column=typeCol).value)
                            else:
                                val = ''
                        elif (receCol != -1):
                            val = str(sheet.cell(row=i, column=receCol + 1).value)
                        else:
                            val = ''
                        self.grid.SetCellValue(rIdx, 3, val)

                        # ///////////////
                        if (sentCol == -2):
                            if (type == 'Sent'):
                                val = str(sheet.cell(row=i, column=typeCol).value)
                            else:
                                val = ''
                        elif (sentCol != -1):
                            val = str(sheet.cell(row=i, column=sentCol + 1).value)
                        else:
                            val = ''
                        self.grid.SetCellValue(rIdx, 4, val)

                        # ///////////////
                        val = str(sheet.cell(row=i, column=descCol + 1).value)
                        self.grid.SetCellValue(rIdx, 6, val)

                        # self.grid.SetCellValue(rIdx, 1, sheet.cell(row=i, column=2).value)
                        # self.grid.SetCellValue(rIdx, 2, sheet.cell(row=i, column=3).value)
                        # self.grid.SetCellValue(rIdx, 3, sheet.cell(row=i, column=4).value)
                        # self.grid.SetCellValue(rIdx, 4, sheet.cell(row=i, column=5).value)

                        rIdx    += 1

                elif (extension == '.xls'):
                    wb      = xlrd.open_workbook(fileName)
                    sheet   = wb.sheet_by_index(0)
                    rCnt    = sheet.nrows
                    for i in range(1, rCnt):
                        self.grid.InsertRows(rIdx, 1)
                        self.grid.SetCellValue(rIdx, 0, org)
                        val = sheet.cell(rowx=i, colx=dateCol).value
                        print(val)
                        val = self.xldate_to_datetime(val)
                        val = str(val)
                        self.grid.SetCellValue(rIdx, 1, val)
                        # ///////////////
                        if (curCol != -1):
                            val = str(sheet.cell(rowx=i, colx=curCol).value)
                        else:
                            val = ''
                        self.grid.SetCellValue(rIdx, 2, val)

                        # ///////////////
                        if (typeCol != -1):
                            type = val = str(sheet.cell(rowx=i, colx=typeCol).value)
                        else:
                            type = val = ''
                        self.grid.SetCellValue(rIdx, 5, val)

                        # ///////////////
                        if (receCol == -2):
                            if (type == 'Received'):
                                val = str(sheet.cell(rowx=i, colx=typeCol - 1).value)
                            else:
                                val = ''
                        elif (receCol != -1):
                            val = str(sheet.cell(rowx=i, colx=receCol).value)
                        else:
                            val = ''
                        self.grid.SetCellValue(rIdx, 3, val)

                        # ///////////////
                        if (sentCol == -2):
                            if (type == 'Sent'):
                                val = str(sheet.cell(rowx=i, colx=typeCol - 1).value)
                            else:
                                val = ''
                        elif (sentCol != -1):
                            val = str(sheet.cell(rowx=i, colx=sentCol).value)
                        else:
                            val = ''
                        self.grid.SetCellValue(rIdx, 4, val)

                        # ///////////////
                        val = str(sheet.cell(rowx=i, colx=descCol).value)
                        self.grid.SetCellValue(rIdx, 6, val)

                        rIdx    += 1

                else:
                    with open(fileName, 'r') as csvFile:
                        reader = csv.reader(csvFile)
                        rCnt = -1
                        for row in reader:
                            rCnt    += 1
                            if (rCnt     == 0):
                                continue

                            print(row)
                            self.grid.InsertRows(rIdx, 1)
                            self.grid.SetCellValue(rIdx, 0, org)

                            datetime_object = datetime.datetime.strptime(row[dateCol], '%m/%d/%Y %H:%M')
                            self.grid.SetCellValue(rIdx, 1, str(datetime_object))
                            self.grid.SetCellValue(rIdx, 2, row[curCol])
                            if (typeCol != -1):
                                self.grid.SetCellValue(rIdx, 5, row[typeCol])
                            else:
                                self.grid.SetCellValue(rIdx, 5, '')

                            if (receCol == -2):
                                if (row[typeCol] == 'Received'):
                                    self.grid.SetCellValue(rIdx, 3, row[typeCol - 1])
                                else:
                                    self.grid.SetCellValue(rIdx, 3, '')
                            elif (receCol != -1):
                                self.grid.SetCellValue(rIdx, 3, row[receCol])
                            else:
                                self.grid.SetCellValue(rIdx, 3, '')

                            if (sentCol == -2):
                                if (row[typeCol] == 'Sent'):
                                    self.grid.SetCellValue(rIdx, 4, row[typeCol - 1])
                                else:
                                    self.grid.SetCellValue(rIdx, 4, '')
                            elif (sentCol != -1):
                                self.grid.SetCellValue(rIdx, 4, row[sentCol])
                            else:
                                self.grid.SetCellValue(rIdx, 4, '')

                            if (descCol != -1):
                                self.grid.SetCellValue(rIdx, 6, row[descCol])
                            else:
                                self.grid.SetCellValue(rIdx, 6, row[descCol])

                            rIdx    += 1
        except:
            wx.MessageBox("An error is occured. ", "Ops", wx.OK | wx.ICON_WARNING)
            return

        self.cmbSort.SetSelection(-1)

    # def on_btnSort_clicked(self, evetn):
    #     self.rows    = []
    #
    #     rCnt    = self.grid.GetNumberRows()
    #     for r in range(rCnt):
    #         self.row     = [self.grid.GetCellValue(r, 0), self.grid.GetCellValue(r, 1), self.grid.GetCellValue(r, 2), \
    #                    self.grid.GetCellValue(r, 3), self.grid.GetCellValue(r, 4), self.grid.GetCellValue(r, 5)]
    #         self.rows.insert(r, self.row)
    #     self.rows.sort(key=self.sort_field)
    #     rCnt = self.grid.GetNumberRows()
    #     for r in range(rCnt):
    #         self.grid.SetCellValue(r, 0, self.rows[r][0])
    #         self.grid.SetCellValue(r, 1, self.rows[r][1])
    #         self.grid.SetCellValue(r, 2, self.rows[r][2])
    #         self.grid.SetCellValue(r, 3, self.rows[r][3])
    #         self.grid.SetCellValue(r, 4, self.rows[r][4])
    #         self.grid.SetCellValue(r, 5, self.rows[r][5])

    def on_cmbSort_item_selected(self, event):
        self.rows    = []

        rCnt    = self.grid.GetNumberRows()
        for r in range(rCnt):
            self.row     = [self.grid.GetCellValue(r, 0), self.grid.GetCellValue(r, 1), self.grid.GetCellValue(r, 2), \
                       self.grid.GetCellValue(r, 3), self.grid.GetCellValue(r, 4), self.grid.GetCellValue(r, 5)]
            self.rows.insert(r, self.row)

        self.sortIndex  = int(self.cmbSort.GetSelection() / 2)
        sortReverse     = self.cmbSort.GetSelection() % 2

        self.rows.sort(key=self.sort_field, reverse=sortReverse)
        rCnt = self.grid.GetNumberRows()
        for r in range(rCnt):
            self.grid.SetCellValue(r, 0, self.rows[r][0])
            self.grid.SetCellValue(r, 1, self.rows[r][1])
            self.grid.SetCellValue(r, 2, self.rows[r][2])
            self.grid.SetCellValue(r, 3, self.rows[r][3])
            self.grid.SetCellValue(r, 4, self.rows[r][4])
            self.grid.SetCellValue(r, 5, self.rows[r][5])

    def on_btnExport_clicked(self, event):
        print("Export")
        with wx.FileDialog(self, "Open XYZ file", wildcard="Comma-Separated Values Files (*.csv)|*.csv",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return ''

            fileName    = fileDialog.GetPath()
            if (fileName == ""):
                return

        self.rows    = []

        rCnt    = self.grid.GetNumberRows()
        for r in range(rCnt):
            self.row     = [self.grid.GetCellValue(r, 0), self.grid.GetCellValue(r, 1), self.grid.GetCellValue(r, 2), \
                       self.grid.GetCellValue(r, 3), self.grid.GetCellValue(r, 4), self.grid.GetCellValue(r, 5)]
            self.rows.insert(r, self.row)

        self.rows.insert(0, ['Date', 'Currency', 'Received', 'Sent', 'Type', 'Description'])
        try:
            with open(fileName, 'w', newline='') as writeFile:
                writer = csv.writer(writeFile)
                writer.writerows(self.rows)
        except:
            wx.MessageBox("Permission is denied.\nClose another applications using output file.", "Ops", wx.OK | wx.ICON_WARNING)
            return

        wx.MessageBox("Export finished successfully.", "Export",
                      wx.OK | wx.ICON_INFORMATION)

# ex = wx.App()
#
# ex.SetTopWindow(MyWin(None, 'Import CSV'))
# ex.MainLoop()

if __name__ == "__main__":
    app = wx.App(False)
    frame = MyWin(None, 'Import CSV')
    # frame.Show()
    app.MainLoop()